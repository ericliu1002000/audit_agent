from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Dict, Iterable, List, Tuple

from openpyxl import load_workbook

from budget_audit.services.normalization import (
    build_embedding_text,
    normalize_tax_flag,
    normalize_text,
    parse_decimal,
    parse_price_range,
)


STANDARD_HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "material_name": ("材料名称", "材料", "名称"),
    "spec_model": ("规格型号", "规格", "型号"),
    "unit": ("单位",),
    "base_price": ("中准价格", "中准价", "基准价"),
    "price_range": ("区间价格", "区间价", "价格区间"),
    "is_tax_included": ("是否含税", "含税"),
}

VENDOR_HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "material_name": ("材料名称", "材料", "名称"),
    "spec_model": ("规格型号", "规格", "型号"),
    "unit": ("单位",),
    "vendor_price": ("用户报价", "供应商报价", "报价", "单价", "投标报价"),
    "is_tax_included": ("是否含税", "含税"),
}


def _read_workbook(uploaded_file):
    uploaded_file.seek(0)
    data = uploaded_file.read()
    workbook = load_workbook(BytesIO(data), data_only=True)
    return workbook


def _normalize_header(value) -> str:
    return normalize_text(value).replace("（", "(").replace("）", ")")


def _find_header_row(worksheet, aliases: Dict[str, Tuple[str, ...]], required_keys: Iterable[str]):
    alias_map = {
        key: {_normalize_header(alias) for alias in alias_values}
        for key, alias_values in aliases.items()
    }
    required = set(required_keys)

    for row_idx in range(1, min(worksheet.max_row, 30) + 1):
        row_values = [
            _normalize_header(worksheet.cell(row=row_idx, column=col_idx).value)
            for col_idx in range(1, worksheet.max_column + 1)
        ]
        matched_keys = set()
        for key, key_aliases in alias_map.items():
            if any(cell in key_aliases for cell in row_values):
                matched_keys.add(key)
        if required.issubset(matched_keys):
            return row_idx
    raise ValueError("Excel 表头不符合模板要求，请检查列名。")


def _build_header_index(worksheet, header_row: int, aliases: Dict[str, Tuple[str, ...]]) -> Dict[str, int]:
    row_values = {
        _normalize_header(worksheet.cell(row=header_row, column=col_idx).value): col_idx
        for col_idx in range(1, worksheet.max_column + 1)
    }
    index: Dict[str, int] = {}
    for key, alias_values in aliases.items():
        for alias in alias_values:
            col = row_values.get(_normalize_header(alias))
            if col:
                index[key] = col
                break
    return index


def parse_standard_price_excel(
    uploaded_file,
    *,
    region: str,
    publish_month: str,
    default_tax_included: bool = True,
) -> List[Dict]:
    """解析政府标准价格 Excel。"""

    region = normalize_text(region)
    publish_month = normalize_text(publish_month)
    if not region or not publish_month:
        raise ValueError("地区与发布月份不能为空。")

    workbook = _read_workbook(uploaded_file)
    sheet = workbook.worksheets[0]

    required = ("material_name", "spec_model", "unit", "base_price", "price_range")
    header_row = _find_header_row(sheet, STANDARD_HEADER_ALIASES, required)
    header_index = _build_header_index(sheet, header_row, STANDARD_HEADER_ALIASES)

    rows: List[Dict] = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        material_name = normalize_text(
            sheet.cell(row=row_idx, column=header_index["material_name"]).value
        )
        spec_model = normalize_text(
            sheet.cell(row=row_idx, column=header_index["spec_model"]).value
        )
        unit = normalize_text(sheet.cell(row=row_idx, column=header_index["unit"]).value)

        # 空行直接跳过
        if not material_name and not spec_model and not unit:
            continue

        base_price = parse_decimal(
            sheet.cell(row=row_idx, column=header_index["base_price"]).value
        )
        if base_price is None:
            raise ValueError(f"第 {row_idx} 行中准价格无效。")

        price_range_value = sheet.cell(
            row=row_idx, column=header_index["price_range"]
        ).value
        price_low, price_high = parse_price_range(price_range_value)

        tax_col = header_index.get("is_tax_included")
        tax_raw = sheet.cell(row=row_idx, column=tax_col).value if tax_col else None
        is_tax_included = normalize_tax_flag(tax_raw, default=default_tax_included)

        embedding_text = build_embedding_text(
            material_name=material_name,
            spec_model=spec_model,
            unit=unit,
            is_tax_included=is_tax_included,
        )

        rows.append(
            {
                "material_name": material_name,
                "spec_model": spec_model,
                "unit": unit,
                "base_price": Decimal(base_price),
                "price_low": Decimal(price_low) if price_low is not None else None,
                "price_high": Decimal(price_high) if price_high is not None else None,
                "is_tax_included": is_tax_included,
                "publish_month": publish_month,
                "region": region,
                "embedding_text": embedding_text,
            }
        )

    if not rows:
        raise ValueError("Excel 中未解析到有效标准价格数据。")

    return rows


def parse_vendor_quote_excel(uploaded_file) -> List[Dict]:
    """解析用户/供应商报价 Excel。"""

    workbook = _read_workbook(uploaded_file)
    sheet = workbook.worksheets[0]

    required = ("material_name", "spec_model", "unit", "vendor_price")
    header_row = _find_header_row(sheet, VENDOR_HEADER_ALIASES, required)
    header_index = _build_header_index(sheet, header_row, VENDOR_HEADER_ALIASES)

    rows: List[Dict] = []
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        material_name = normalize_text(
            sheet.cell(row=row_idx, column=header_index["material_name"]).value
        )
        spec_model = normalize_text(
            sheet.cell(row=row_idx, column=header_index["spec_model"]).value
        )
        unit = normalize_text(sheet.cell(row=row_idx, column=header_index["unit"]).value)

        if not material_name and not spec_model and not unit:
            continue
        if not material_name:
            raise ValueError(f"第 {row_idx} 行材料名称为空。")

        vendor_price = parse_decimal(
            sheet.cell(row=row_idx, column=header_index["vendor_price"]).value
        )
        tax_col = header_index.get("is_tax_included")
        tax_raw = sheet.cell(row=row_idx, column=tax_col).value if tax_col else None
        is_tax_included = normalize_tax_flag(tax_raw, default=True)

        embedding_text = build_embedding_text(
            material_name=material_name,
            spec_model=spec_model,
            unit=unit,
            is_tax_included=is_tax_included,
        )

        rows.append(
            {
                "row_number": row_idx,
                "material_name": material_name,
                "spec_model": spec_model,
                "unit": unit,
                "vendor_price": Decimal(vendor_price) if vendor_price is not None else None,
                "is_tax_included": is_tax_included,
                "embedding_text": embedding_text,
            }
        )

    if not rows:
        raise ValueError("Excel 中未解析到有效报价数据。")

    return rows

