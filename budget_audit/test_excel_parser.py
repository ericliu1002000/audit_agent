import unittest
from io import BytesIO

from openpyxl import Workbook

from budget_audit.services.excel_parser import (
    parse_standard_price_excel,
    parse_vendor_quote_excel,
)


def _build_standard_workbook_bytes():
    wb = Workbook()
    ws = wb.active
    ws.append(["2025年12月天津市建设工程主要材料市场价格（含税价格）"])
    ws.append([])
    ws.append(["序号", "材料名称", "规格型号", "单位", "中准价格", "区间价格", "是否含税"])
    ws.append([1, "矿渣硅酸盐水泥", "32.5级 散装", "t", 382.22, "285.80-515.00", "是"])
    ws.append([2, "普通硅酸盐水泥", "42.5级 袋装", "t", 458.44, "379.76-590.00", "是"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_vendor_workbook_bytes():
    wb = Workbook()
    ws = wb.active
    ws.append(["材料名称", "规格型号", "单位", "用户报价", "是否含税"])
    ws.append(["32.5级矿渣水泥", "32.5级 散装", "t", 390.0, "是"])
    ws.append(["普通硅酸盐水泥", "42.5级 袋装", "t", 455.0, "是"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ExcelParserTests(unittest.TestCase):
    def test_parse_standard_price_excel_success(self):
        file_obj = _build_standard_workbook_bytes()
        rows = parse_standard_price_excel(
            file_obj, region="天津市", publish_month="2025-12"
        )
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["material_name"], "矿渣硅酸盐水泥")
        self.assertEqual(str(rows[0]["base_price"]), "382.22")
        self.assertEqual(str(rows[0]["price_low"]), "285.80")
        self.assertEqual(str(rows[0]["price_high"]), "515.00")
        self.assertTrue(rows[0]["is_tax_included"])
        self.assertIn("规格型号:32.5级 散装", rows[0]["embedding_text"])

    def test_parse_vendor_quote_excel_success(self):
        file_obj = _build_vendor_workbook_bytes()
        rows = parse_vendor_quote_excel(file_obj)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["material_name"], "32.5级矿渣水泥")
        self.assertEqual(str(rows[0]["vendor_price"]), "390")
        self.assertTrue(rows[0]["is_tax_included"])

    def test_parse_vendor_quote_excel_missing_required_header(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["材料名称", "规格型号", "单位"])
        ws.append(["矿渣硅酸盐水泥", "32.5级 散装", "t"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        with self.assertRaises(ValueError) as ctx:
            parse_vendor_quote_excel(buf)

        self.assertIn("表头不符合模板要求", str(ctx.exception))


if __name__ == "__main__":
    unittest.main()

