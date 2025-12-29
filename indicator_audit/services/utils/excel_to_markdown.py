"""
parse_indicator_excel
----------------------
本模块用于将财政绩效指标申报表（Excel）转换成扁平化的 Markdown 表格文本。
在 LLM 参与的审核流程中，扁平的 Markdown 更容易被模型读取与理解，
因此需要一个专门的工具方法来处理原始 Excel 中复杂的合并单元格以及格式差异。
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import List

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def clean_text(value) -> str:
    """
    标准化 Excel 单元格文本：去除首尾及内部空白，以便 LLM 正确识别字段。
    业务表格常见“绩 效 目 标”这种分散对齐写法，如果不清洗，会被当成多个 token。
    """

    if value is None:
        return ""
    text = str(value).strip()
    # 使用正则移除字符串内部所有空白字符，避免中文“分散对齐”留下多余空格。
    return re.sub(r"\s+", "", text)


def parse_excel_to_markdown(file_path: str) -> str:
    """
    将 Excel 文件解析为 Markdown 表格字符串（针对指标审核场景做了定制优化）。
    处理流程包括：
    1. 基于 sheet 尺寸初始化 Grid，填充原始单元格值；
    2. 展开合并单元格，将左上角值覆盖到区域内所有格子；
    3. 使用 clean_text 进行文本标准化，按原始网格输出 Markdown 表格。

    参数:
        file_path (str): Excel 文件的绝对路径或相对路径。

    返回:
        str: 解析后的 Markdown 表格内容。

    异常:
        ValueError: 当上传的并非 xlsx 文件，或 openpyxl 无法正确解析该文件时抛出。
    """

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {path}")

    if path.suffix.lower() != ".xlsx":
        raise ValueError("请上传xlsx格式的表格文件")

    # data_only=True 会读取单元格计算后的值，避免公式字符串干扰规则校验。
    try:
        workbook = load_workbook(filename=path, data_only=True)
    except (InvalidFileException, ValueError) as exc:
        # 当 openpyxl 无法解析该文件时，通常意味着格式不符合预期，这里抛出统一的中文提示。
        raise ValueError("请上传xlsx格式的表格文件") from exc
    sheet = workbook.worksheets[0]

    max_rows = sheet.max_row or 0
    max_cols = sheet.max_column or 0
    if max_rows == 0 or max_cols == 0:
        return ""

    # Step 1: 初始化 Grid 并填充原始值
    grid: List[List[object]] = []
    for row_index in range(1, max_rows + 1):
        row_values = []
        for col_index in range(1, max_cols + 1):
            row_values.append(sheet.cell(row=row_index, column=col_index).value)
        grid.append(row_values)

    # Step 2: 处理合并单元格，填充合并区域内所有单元格
    for merged_range in list(sheet.merged_cells.ranges):
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        for row_index in range(min_row, max_row + 1):
            row_values = grid[row_index - 1]
            for col_index in range(min_col, max_col + 1):
                row_values[col_index - 1] = top_left_value

    # Step 3: 清洗文本并记录关键字段
    required_keywords = {"一级指标", "二级指标", "三级指标"}
    found_keywords = set()
    final_rows: List[List[str]] = []
    for row in grid:
        cleaned_row = [clean_text(cell) for cell in row]
        for text in cleaned_row:
            if text in required_keywords:
                found_keywords.add(text)
        final_rows.append(cleaned_row)

    if required_keywords - found_keywords:
        raise ValueError("模板格式错误，缺少指标数据（确认文件包含一级指标、二级指标、三级指标）")

    if not any(any(cell != "" for cell in row) for row in final_rows):
        return ""

    markdown_lines = ["|" + "|".join(row) + "|" for row in final_rows]

    # Markdown 表格需要在首行（表头）后插入分割线，用于区分表头和主体。
    divider = "|" + "|".join(["---"] * max_cols) + "|"
    markdown_lines.insert(1, divider)

    return "\n".join(markdown_lines)
