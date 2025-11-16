from __future__ import annotations

from io import BytesIO
from typing import Dict, Iterable, List, Tuple

from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from urllib.parse import quote

from indicators.models import FundUsage, Indicator
from regions.models import Province

EXPORT_HEADERS = [
    "编码",
    "资金用途",
    "一级指标",
    "二级指标",
    "三级指标",
    "指标性质",
    "计量单位",
    "指标解释",
    "省份",
]

IMPORT_HEADERS = [
    "编码",
    "资金用途",
    "一级指标",
    "二级指标",
    "三级指标",
    "指标性质",
    "计量单位",
    "指标解释",
    "省份",
]


def export_indicators_excel(queryset, province_name: str = "全部省份") -> HttpResponse:
    """导出已筛选的指标为 Excel."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "指标列表"
    worksheet.append(EXPORT_HEADERS)

    sorted_queryset = queryset.order_by(
        "fund_usage__name",
        "level_1",
        "level_2",
        "level_3",
    )

    for indicator in sorted_queryset:
        province_name_value = (
            indicator.province_id.name if indicator.province_id else ""
        )
        fund_usage_name = indicator.fund_usage.name if indicator.fund_usage else ""
        worksheet.append(
            [
                indicator.business_code or "",
                fund_usage_name,
                indicator.level_1,
                indicator.level_2,
                indicator.level_3,
                indicator.nature,
                indicator.unit,
                indicator.explanation or "",
                province_name_value,
            ]
        )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)

    timestamp = timezone.now().strftime("%y%m%d-%H%M%S")
    filename = f"{province_name}指标体系-{timestamp}.xlsx"
    response = HttpResponse(
        stream.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    safe_filename = quote(filename)
    response["Content-Disposition"] = (
        f'attachment; filename="export.xlsx"; filename*=UTF-8\'\'{safe_filename}'
    )
    return response


def full_sync_from_excel(uploaded_file, source_tag: str | None = None) -> Dict[str, int]:
    """读取 Excel 并对指定 source_tag 的指标执行全量同步（含软删除）。"""

    normalized_source_tag = (source_tag or "").strip()

    uploaded_file.seek(0)
    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active

    header_row = next(
        worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None
    )
    if not header_row:
        raise ValidationError("Excel 表头不能为空")

    header_map = {
        str(value).strip(): idx
        for idx, value in enumerate(header_row)
        if value is not None
    }
    missing_headers = [header for header in IMPORT_HEADERS if header not in header_map]
    if missing_headers:
        raise ValidationError(f"缺少必要的列: {', '.join(missing_headers)}")

    excel_data_map: Dict[
        Tuple[str, str, str], Dict[str, str]
    ] = {}
    province_names: set[str] = set()
    fund_usage_names: set[str] = set()

    def _clean_value(value) -> str:
        """strip 所有空白字符."""

        if value is None:
            return ""
        return "".join(str(value).split())

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        def _get_value(column_name: str) -> str:
            value = row[header_map[column_name]]
            return _clean_value(value)

        business_code = _get_value("编码")
        province_name = _get_value("省份")
        fund_usage_name = _get_value("资金用途")
        level_1 = _get_value("一级指标")
        level_2 = _get_value("二级指标")
        level_3 = _get_value("三级指标")
        nature = _get_value("指标性质")
        unit = _get_value("计量单位")
        explanation_value = _clean_value(row[header_map["指标解释"]])

        if not province_name:
            raise ValidationError("省份不能为空")
        if not fund_usage_name or not level_3:
            continue

        key = (province_name, fund_usage_name, level_3)
        excel_data_map[key] = {
            "province_name": province_name,
            "fund_usage_name": fund_usage_name,
            "business_code": business_code,
            "level_1": level_1,
            "level_2": level_2,
            "level_3": level_3,
            "nature": nature,
            "unit": unit,
            "explanation": explanation_value,
        }
        province_names.add(province_name)
        fund_usage_names.add(fund_usage_name)

    if not excel_data_map:
        raise ValidationError("Excel 中没有可同步的数据")

    provinces = Province.objects.filter(name__in=province_names)
    province_cache = {province.name: province for province in provinces}
    missing_provinces = province_names - province_cache.keys()
    if missing_provinces:
        raise ValidationError(f"系统中缺少这些省份: {', '.join(sorted(missing_provinces))}")

    fund_usage_cache: Dict[str, FundUsage] = {
        fu.name: fu for fu in FundUsage.objects.filter(name__in=fund_usage_names)
    }
    for fund_usage_name in fund_usage_names:
        if fund_usage_name not in fund_usage_cache:
            fund_usage_cache[fund_usage_name], _ = FundUsage.objects.get_or_create(
                name=fund_usage_name
            )

    excel_keys = set(excel_data_map.keys())
    db_keys = set(
        Indicator.all_objects.filter(
            is_active=True, province_id__name__in=province_names
        ).values_list("province_id__name", "fund_usage__name", "level_3")
    )

    existing_indicators = Indicator.all_objects.filter(
        province_id__name__in=province_names
    ).select_related("province_id", "fund_usage")
    existing_map = {
        (indicator.province_id.name, indicator.fund_usage.name, indicator.level_3): indicator
        for indicator in existing_indicators
    }
    existing_keys = set(existing_map.keys())

    keys_to_create = excel_keys - db_keys
    keys_to_delete = db_keys - excel_keys

    inactive_matching_keys = keys_to_create & existing_keys
    keys_to_create -= inactive_matching_keys

    keys_to_update = (excel_keys & db_keys) | inactive_matching_keys

    results = {"created": 0, "updated": 0, "soft_deleted": 0}

    with transaction.atomic():
        new_indicators: List[Indicator] = []
        for key in keys_to_create:
            data = excel_data_map[key]
            new_indicators.append(
                Indicator(
                    business_code=data["business_code"],
                    fund_usage=fund_usage_cache[data["fund_usage_name"]],
                    level_1=data["level_1"],
                    level_2=data["level_2"],
                    level_3=data["level_3"],
                    nature=data["nature"],
                    unit=data["unit"],
                    explanation=data["explanation"],
                    province_id=province_cache[data["province_name"]],
                    source_tag=normalized_source_tag,
                    is_active=True,
                    is_vectorized=False,
                )
            )
        if new_indicators:
            Indicator.objects.bulk_create(new_indicators)
            results["created"] = len(new_indicators)



        update_objects: List[Indicator] = []
        for key in keys_to_update:
            indicator = existing_map[key]
            data = excel_data_map[key]
            new_fund_usage = fund_usage_cache[data["fund_usage_name"]]
            new_province = province_cache[data["province_name"]]

            should_update = False
            if (indicator.business_code or "") != (data["business_code"] or ""):
                should_update = True
            if indicator.fund_usage_id != new_fund_usage.id:
                should_update = True
            if indicator.province_id_id != new_province.id:
                should_update = True
            if (indicator.level_1 or "") != (data["level_1"] or ""):
                should_update = True
            if (indicator.level_2 or "") != (data["level_2"] or ""):
                should_update = True
            if (indicator.level_3 or "") != (data["level_3"] or ""):
                should_update = True
            if (indicator.nature or "") != (data["nature"] or ""):
                should_update = True
            if (indicator.unit or "") != (data["unit"] or ""):
                should_update = True
            if (indicator.explanation or "") != (data["explanation"] or ""):
                should_update = True
            # if (indicator.source_tag or "") != normalized_source_tag:
            #     should_update = True
            if not indicator.is_active:
                should_update = True

            if not should_update:
                continue

            indicator.business_code = data["business_code"]
            indicator.fund_usage = new_fund_usage
            indicator.province_id = new_province
            indicator.level_1 = data["level_1"]
            indicator.level_2 = data["level_2"]
            indicator.level_3 = data["level_3"]
            indicator.nature = data["nature"]
            indicator.unit = data["unit"]
            indicator.explanation = data["explanation"]
            indicator.source_tag = normalized_source_tag
            indicator.is_active = True
            indicator.is_vectorized = False
            update_objects.append(indicator)


        if update_objects:
            Indicator.all_objects.bulk_update(
                update_objects,
                [
                    "business_code",
                    "fund_usage",
                    "province_id",
                    "level_1",
                    "level_2",
                    "level_3",
                    "nature",
                    "unit",
                    "explanation",
                    "source_tag",
                    "is_active",
                    "is_vectorized",
                ],
            )
            results["updated"] = len(update_objects)

        if keys_to_delete:
            results["soft_deleted"] = _soft_delete_by_keys(keys_to_delete)

    return results


def _soft_delete_by_keys(keys: Iterable[Tuple[str, str, str]]) -> int:
    """批量软删除给定 key 所对应的指标。"""

    total_deleted = 0
    key_list = list(keys)
    chunk_size = 50
    for start in range(0, len(key_list), chunk_size):
        chunk = key_list[start : start + chunk_size]
        condition = Q()
        for province_name, fund_usage_name, level_3 in chunk:
            condition |= Q(
                province_id__name=province_name,
                fund_usage__name=fund_usage_name,
                level_3=level_3,
                is_active=True,
            )
        if condition:
            total_deleted += Indicator.all_objects.filter(condition).update(
                is_active=False
            )
    return total_deleted
