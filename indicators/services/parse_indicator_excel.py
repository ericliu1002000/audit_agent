"""
parse_indicator_excel
----------------------
本模块用于将财政绩效指标申报表（Excel）转换成扁平化的 Markdown 表格文本。
在 LLM 参与的审核流程中，扁平的 Markdown 更容易被模型读取与理解，
因此需要一个专门的工具方法来处理原始 Excel 中复杂的合并单元格以及格式差异。
"""

from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def parse_excel_to_markdown(file_path: str) -> str:
    """
    将 Excel 文件解析为 Markdown 表格字符串。   这里专门针对指标模型进行优化。并不具备通用性。

    参数:
        file_path (str): Excel 文件的绝对路径或相对路径。

    返回:
        str: 解析后的 Markdown 表格内容。

    异常:
        ValueError: 当上传的并非 xlsx 文件，或 openpyxl 无法正确解析该文件时抛出。

    示例:
        >>> markdown = parse_excel_to_markdown("indicators/doc/indicator-example.xlsx")
        >>> print(markdown)
    """

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {path}")

    if path.suffix.lower() != ".xlsx":
        raise ValueError("请上传xlsx格式的表格文件")

    # data_only=True 会读取单元格计算后的值，避免公式字符串干扰规则校验。
    try:
        workbook = load_workbook(filename=path, data_only=True)
    except (InvalidFileException, ValueError) as exc:
        # 当 openpyxl 无法解析该文件时，通常意味着格式不符合预期，这里抛出统一的中文提示。
        raise ValueError("请上传xlsx格式的表格文件") from exc
    sheet = workbook.worksheets[0]

    # 先复制 merged_cells.ranges，再遍历，避免在 unmerge 过程中修改原列表导致的迭代问题。
    for merged_range in list(sheet.merged_cells.ranges):
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        # 合并区域只有左上角有值，这里需要把左上角的值填充到整个区域，确保扁平化后信息不丢失。
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(merged_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).value = top_left_value

    markdown_lines: List[str] = []
    for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
        original_row = ["" if cell is None else str(cell).strip() for cell in row]

        if row_index < 10:
            # 仅对前 10 行做行内去重，避免指标列表等长表格信息被过度清洗。
            cleaned_row: List[str] = []
            for col_index, value in enumerate(original_row):
                if col_index == 0:
                    cleaned_row.append(value)
                    continue
                # 必须以“原始值”做对比，不能一边修改一边比较，否则会出现 A->"" 后再把后面的 A 误认为不相同。
                prev_original_value = original_row[col_index - 1]
                if value == prev_original_value:
                    cleaned_row.append("")
                else:
                    cleaned_row.append(value)
        else:
            cleaned_row = original_row

        if all(cell == "" for cell in cleaned_row):
            continue
        markdown_lines.append("|" + "|".join(cleaned_row) + "|")

    if not markdown_lines:
        return ""

    # Markdown 表格需要在首行（表头）后插入分割线，用于区分表头和主体。
    header_cell_count = markdown_lines[0].count("|") - 1
    divider = "|" + "|".join(["---"] * header_cell_count) + "|"
    markdown_lines.insert(1, divider)

    return "\n".join(markdown_lines)


if __name__ == "__main__":
    # 使用项目中提供的示例 Excel 做演示（需确保为 xlsx 格式），可根据需要替换路径。
    # example_path = "indicators/doc/indicator-example.xlsx"
    example_path = "/Users/liuxiaoqi/SynologyDrive/work/势术/合作/审计智能体/指标相关/实例/天津-无线电管理设施运维.xlsx"
    try:
        print(parse_excel_to_markdown(example_path))
    except Exception as exc:
        print(f"解析 Excel 失败: {exc}")
