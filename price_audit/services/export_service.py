"""价格审核表导出服务。"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from price_audit.models import PriceAuditSubmission


def _to_excel_number(value: Decimal | None):
    if value is None:
        return None
    return float(value)


def build_audited_excel_content(submission: PriceAuditSubmission) -> bytes:
    """根据审核结果回填审核表 Excel。"""

    workbook = load_workbook(submission.source_file.path)
    sheet = workbook.worksheets[0]

    sheet.cell(row=1, column=9).value = sheet.cell(row=1, column=9).value or "审核"
    sheet.cell(row=1, column=15).value = sheet.cell(row=1, column=15).value or "审减原因/未审减原因"
    sheet.cell(row=2, column=9).value = sheet.cell(row=2, column=9).value or "计量单位"
    sheet.cell(row=2, column=10).value = sheet.cell(row=2, column=10).value or "单价（元）"
    sheet.cell(row=2, column=11).value = sheet.cell(row=2, column=11).value or "数量"
    sheet.cell(row=2, column=12).value = sheet.cell(row=2, column=12).value or "天数"
    sheet.cell(row=2, column=13).value = sheet.cell(row=2, column=13).value or "审核金额（元）"
    sheet.cell(row=2, column=14).value = sheet.cell(row=2, column=14).value or "审减金额（元）"

    for row in submission.rows.select_related("decision").order_by("excel_row_no"):
        decision = getattr(row, "decision", None)
        if decision is None:
            continue
        excel_row_no = row.excel_row_no
        sheet.cell(row=excel_row_no, column=9).value = decision.reviewed_unit
        sheet.cell(row=excel_row_no, column=10).value = _to_excel_number(decision.reviewed_unit_price)
        sheet.cell(row=excel_row_no, column=11).value = _to_excel_number(decision.reviewed_quantity)
        sheet.cell(row=excel_row_no, column=12).value = _to_excel_number(decision.reviewed_days)
        sheet.cell(row=excel_row_no, column=13).value = _to_excel_number(decision.reviewed_amount)
        sheet.cell(row=excel_row_no, column=14).value = _to_excel_number(decision.reduction_amount)
        sheet.cell(row=excel_row_no, column=15).value = decision.reason

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
