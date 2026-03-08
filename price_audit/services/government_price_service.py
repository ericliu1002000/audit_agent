"""政府标准价导入与模板 service。"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from django.core.files.base import File
from django.db import transaction
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from price_audit.models import GovernmentPriceBatch, GovernmentPriceItem
from price_audit.tasks import dispatch_vectorize_government_price_batch
from price_audit.services.normalization import (
    build_embedding_text,
    normalize_tax_flag,
    normalize_text,
    normalize_text_no_space,
    parse_decimal,
    parse_price_range,
)


TEMPLATE_HEADERS = (
    "材料名称",
    "规格型号",
    "单位",
    "中准价格",
    "区间最低价",
    "区间最高价",
    "说明",
    "是否含税",
)

HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "material_name": ("材料名称", "材料", "名称"),
    "spec_model": ("规格型号", "规格", "型号"),
    "unit": ("单位",),
    "benchmark_price": ("中准价格", "中准价", "基准价"),
    "price_low": ("区间最低价", "最低价"),
    "price_high": ("区间最高价", "最高价"),
    "price_range": ("区间价格", "区间价", "价格区间"),
    "description": ("说明", "备注"),
    "is_tax_included": ("是否含税", "含税"),
}


@dataclass(frozen=True)
class ParsedGovernmentPriceRow:
    """Excel 中解析出的一条政府标准价记录。"""

    row_no: int
    material_name_raw: str
    material_name_normalized: str
    spec_model_raw: str
    spec_model_normalized: str
    unit_raw: str
    unit_normalized: str
    benchmark_price: Any
    price_min: Any | None
    price_max: Any | None
    description: str
    is_tax_included: bool
    raw_row_data: dict[str, Any]


@dataclass(frozen=True)
class GovernmentPriceImportResult:
    """一次导入完成后的结果摘要。"""

    batch: GovernmentPriceBatch
    parsed_rows: int
    created_rows: int
    updated_rows: int
    deleted_rows: int
    vector_task_dispatched: bool


class GovernmentPriceService:
    """政府标准价模板、解析和导入的唯一事实来源。"""

    def build_template_content(self) -> bytes:
        """
        功能说明:
            生成后台可下载的政府标准价 Excel 模板，统一列名和填写口径。
        使用示例:
            content = government_price_service.build_template_content()
        输出参数:
            bytes: `.xlsx` 文件的二进制内容，可直接写入 HTTP 响应。
        """

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "政府标准价"
        sheet.append(TEMPLATE_HEADERS)
        sheet.append(
            [
                "矿渣硅酸盐水泥",
                "32.5级 散装",
                "t",
                "379.42",
                "285.80",
                "515.00",
                "示例数据，可删除。",
                "是",
            ]
        )

        help_sheet = workbook.create_sheet("填写说明")
        help_sheet.append(["说明项", "内容"])
        help_sheet.append(["年份", "在后台上传表单中填写，不需要放进 Excel。"])
        help_sheet.append(["地区", "在后台上传表单中填写，不需要放进 Excel。"])
        help_sheet.append(["区间价格", "可选填写；推荐拆成“区间最低价”和“区间最高价”两列。"])
        help_sheet.append(["是否含税", "若列缺失或单元格为空，系统默认按含税处理。"])
        help_sheet.append(["单位换算", "当前版本不做程序换算，请尽量保持官方标准价单位清晰。"])

        for sheet_name in workbook.worksheets:
            for column in sheet_name.columns:
                sheet_name.column_dimensions[column[0].column_letter].width = 18

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def parse_excel(
        self,
        uploaded_file,
        *,
        default_tax_included: bool = True,
    ) -> list[ParsedGovernmentPriceRow]:
        """
        功能说明:
            解析管理员上传的政府标准价 Excel，并转换成结构化行对象列表。
            该过程只负责解析和校验，不负责入库。
        使用示例:
            rows = government_price_service.parse_excel(uploaded_file)
        输入参数:
            uploaded_file: Django 上传文件对象。
            default_tax_included: 当表格缺少“是否含税”列时是否默认按含税处理。
        输出参数:
            list[ParsedGovernmentPriceRow]: 可直接用于入库的结构化明细行列表。
        """

        workbook = self._read_workbook(uploaded_file)
        sheet = workbook.worksheets[0]

        header_row = self._find_header_row(sheet)
        header_index = self._build_header_index(sheet, header_row)

        rows: list[ParsedGovernmentPriceRow] = []
        for row_no in range(header_row + 1, sheet.max_row + 1):
            material_name_raw = normalize_text(
                sheet.cell(row=row_no, column=header_index["material_name"]).value
            )
            spec_model_raw = normalize_text(
                sheet.cell(row=row_no, column=header_index["spec_model"]).value
            )
            unit_raw = normalize_text(
                sheet.cell(row=row_no, column=header_index["unit"]).value
            )

            if not material_name_raw and not spec_model_raw and not unit_raw:
                continue
            if not material_name_raw:
                raise ValueError(f"第 {row_no} 行材料名称为空。")

            benchmark_price = parse_decimal(
                sheet.cell(row=row_no, column=header_index["benchmark_price"]).value
            )
            if benchmark_price is None:
                raise ValueError(f"第 {row_no} 行中准价格无效。")

            price_min, price_max = self._parse_price_bounds(
                sheet,
                row_no=row_no,
                header_index=header_index,
            )
            if (
                price_min is not None
                and price_max is not None
                and price_min > price_max
            ):
                raise ValueError(f"第 {row_no} 行区间最低价不能大于区间最高价。")

            description = normalize_text(
                sheet.cell(
                    row=row_no,
                    column=header_index["description"],
                ).value
            ) if header_index.get("description") else ""
            tax_col = header_index.get("is_tax_included")
            tax_raw = sheet.cell(row=row_no, column=tax_col).value if tax_col else None
            is_tax_included = normalize_tax_flag(
                tax_raw,
                default=default_tax_included,
            )

            rows.append(
                ParsedGovernmentPriceRow(
                    row_no=row_no,
                    material_name_raw=material_name_raw,
                    material_name_normalized=normalize_text(material_name_raw),
                    spec_model_raw=spec_model_raw,
                    spec_model_normalized=normalize_text_no_space(spec_model_raw),
                    unit_raw=unit_raw,
                    unit_normalized=normalize_text_no_space(unit_raw),
                    benchmark_price=benchmark_price,
                    price_min=price_min,
                    price_max=price_max,
                    description=description,
                    is_tax_included=is_tax_included,
                    raw_row_data={
                        "material_name": material_name_raw,
                        "spec_model": spec_model_raw,
                        "unit": unit_raw,
                        "benchmark_price": str(benchmark_price),
                        "price_min": str(price_min) if price_min is not None else None,
                        "price_max": str(price_max) if price_max is not None else None,
                        "description": description,
                        "is_tax_included": is_tax_included,
                    },
                )
            )

        if not rows:
            raise ValueError("Excel 中未解析到有效政府标准价数据。")
        return rows

    def import_excel(
        self,
        uploaded_file,
        *,
        region_name: str,
        year: int,
        uploaded_by=None,
        remark: str = "",
        default_tax_included: bool = True,
    ) -> GovernmentPriceImportResult:
        """
        功能说明:
            导入一份政府标准价 Excel，并在同地区同年份的当前批次上执行增量同步。
        使用示例:
            result = government_price_service.import_excel(
                uploaded_file,
                region_name="天津",
                year=2026,
                uploaded_by=request.user,
            )
        输入参数:
            uploaded_file: Django 上传文件对象。
            region_name: 本次导入所属地区。
            year: 本次导入所属年份。
            uploaded_by: 当前上传用户，可为空。
            remark: 后台备注。
            default_tax_included: 当表格缺少“是否含税”列时是否默认按含税处理。
        输出参数:
            GovernmentPriceImportResult: 导入后的批次信息，以及新增/更新/删除摘要。
        """

        region_name = normalize_text(region_name)
        if not region_name:
            raise ValueError("地区不能为空。")
        if not year:
            raise ValueError("年份不能为空。")

        parsed_rows = self.parse_excel(
            uploaded_file,
            default_tax_included=default_tax_included,
        )
        excel_row_map = {self._build_row_key_from_row(row): row for row in parsed_rows}
        if not excel_row_map:
            raise ValueError("Excel 中未解析到有效政府标准价数据。")

        with transaction.atomic():
            existing_batches = list(
                GovernmentPriceBatch.objects.select_for_update()
                .filter(region_name=region_name, year=year)
                .order_by("-is_active", "-created_at")
            )
            batch = existing_batches[0] if existing_batches else None
            extra_batch_ids = [item.id for item in existing_batches[1:]]
            if extra_batch_ids:
                GovernmentPriceBatch.objects.filter(id__in=extra_batch_ids).update(
                    is_active=False,
                    deactivated_at=timezone.now(),
                )

            if batch is None:
                batch = GovernmentPriceBatch.objects.create(
                    region_name=region_name,
                    year=year,
                    source_filename=uploaded_file.name or "",
                    uploaded_by=uploaded_by,
                    total_rows=len(parsed_rows),
                    success_rows=len(parsed_rows),
                    vector_status=GovernmentPriceBatch.VectorStatus.PENDING,
                    vector_total=0,
                    vector_success=0,
                    vector_failed=0,
                    last_error="",
                    remark=normalize_text(remark),
                    is_active=True,
                    deactivated_at=None,
                )
            else:
                batch.region_name = region_name
                batch.year = year
                batch.source_filename = uploaded_file.name or ""
                batch.uploaded_by = uploaded_by
                batch.total_rows = len(parsed_rows)
                batch.success_rows = len(parsed_rows)
                batch.vector_status = GovernmentPriceBatch.VectorStatus.PENDING
                batch.vector_total = 0
                batch.vector_success = 0
                batch.vector_failed = 0
                batch.vector_task_id = ""
                batch.vector_queued_at = None
                batch.vector_started_at = None
                batch.vectorized_at = None
                batch.last_error = ""
                batch.remark = normalize_text(remark)
                batch.is_active = True
                batch.deactivated_at = None
                batch.save(
                    update_fields=[
                        "region_name",
                        "year",
                        "source_filename",
                        "uploaded_by",
                        "total_rows",
                        "success_rows",
                        "vector_status",
                        "vector_total",
                        "vector_success",
                        "vector_failed",
                        "vector_task_id",
                        "vector_queued_at",
                        "vector_started_at",
                        "vectorized_at",
                        "last_error",
                        "remark",
                        "is_active",
                        "deactivated_at",
                        "updated_at",
                    ]
                )
            uploaded_file.seek(0)
            batch.source_file.save(
                uploaded_file.name,
                File(uploaded_file),
                save=False,
            )
            batch.save(update_fields=["source_file"])

            existing_items = list(batch.items.all())
            existing_item_map = {
                self._build_row_key_from_item(item): item for item in existing_items
            }
            excel_keys = set(excel_row_map.keys())
            existing_keys = set(existing_item_map.keys())

            keys_to_create = excel_keys - existing_keys
            keys_to_delete = existing_keys - excel_keys
            keys_to_check = excel_keys & existing_keys

            created_items = []
            updated_items = []
            deleted_item_ids = [existing_item_map[key].id for key in keys_to_delete]

            if deleted_item_ids:
                GovernmentPriceItem.objects.filter(id__in=deleted_item_ids).delete()

            for key in keys_to_create:
                row = excel_row_map[key]
                embedding_text = self._build_embedding_text_from_row(row)
                created_items.append(
                    GovernmentPriceItem(
                        batch=batch,
                        row_no=row.row_no,
                        material_name_raw=row.material_name_raw,
                        material_name_normalized=row.material_name_normalized,
                        spec_model_raw=row.spec_model_raw,
                        spec_model_normalized=row.spec_model_normalized,
                        unit_raw=row.unit_raw,
                        unit_normalized=row.unit_normalized,
                        benchmark_price=row.benchmark_price,
                        price_min=row.price_min,
                        price_max=row.price_max,
                        description=row.description,
                        is_tax_included=row.is_tax_included,
                        embedding_text=embedding_text,
                        is_vectorized=False,
                        raw_row_data={
                            **row.raw_row_data,
                            "embedding_text": embedding_text,
                        },
                    )
                )

            for key in keys_to_check:
                item = existing_item_map[key]
                row = excel_row_map[key]
                old_embedding_text = item.embedding_text
                new_embedding_text = self._build_embedding_text_from_row(row)
                should_update = False

                updated_fields = {
                    "row_no": row.row_no,
                    "material_name_raw": row.material_name_raw,
                    "material_name_normalized": row.material_name_normalized,
                    "spec_model_raw": row.spec_model_raw,
                    "spec_model_normalized": row.spec_model_normalized,
                    "unit_raw": row.unit_raw,
                    "unit_normalized": row.unit_normalized,
                    "benchmark_price": row.benchmark_price,
                    "price_min": row.price_min,
                    "price_max": row.price_max,
                    "description": row.description,
                    "is_tax_included": row.is_tax_included,
                    "embedding_text": new_embedding_text,
                    "raw_row_data": {
                        **row.raw_row_data,
                        "embedding_text": new_embedding_text,
                    },
                }
                for field_name, new_value in updated_fields.items():
                    if getattr(item, field_name) != new_value:
                        setattr(item, field_name, new_value)
                        should_update = True

                if should_update:
                    if old_embedding_text != new_embedding_text:
                        item.is_vectorized = False
                    updated_items.append(item)

            if created_items:
                GovernmentPriceItem.objects.bulk_create(created_items, batch_size=500)

            if updated_items:
                GovernmentPriceItem.objects.bulk_update(
                    updated_items,
                    [
                        "row_no",
                        "material_name_raw",
                        "material_name_normalized",
                        "spec_model_raw",
                        "spec_model_normalized",
                        "unit_raw",
                        "unit_normalized",
                        "benchmark_price",
                        "price_min",
                        "price_max",
                        "description",
                        "is_tax_included",
                        "embedding_text",
                        "is_vectorized",
                        "raw_row_data",
                    ],
                    batch_size=500,
                )

            needs_vector_sync = bool(
                deleted_item_ids
                or batch.items.filter(is_vectorized=False).exists()
            )
            if needs_vector_sync:
                transaction.on_commit(
                    lambda: dispatch_vectorize_government_price_batch(batch.id, deleted_item_ids)
                )
            else:
                batch.vector_status = GovernmentPriceBatch.VectorStatus.ACTIVE
                batch.vector_total = 0
                batch.vector_success = 0
                batch.vector_failed = 0
                batch.vector_task_id = ""
                batch.vector_queued_at = None
                batch.vector_started_at = None
                batch.vectorized_at = timezone.now()
                batch.last_error = ""
                batch.save(
                    update_fields=[
                        "vector_status",
                        "vector_total",
                        "vector_success",
                        "vector_failed",
                        "vector_task_id",
                        "vector_queued_at",
                        "vector_started_at",
                        "vectorized_at",
                        "last_error",
                        "updated_at",
                    ]
                )

        return GovernmentPriceImportResult(
            batch=batch,
            parsed_rows=len(parsed_rows),
            created_rows=len(created_items),
            updated_rows=len(updated_items),
            deleted_rows=len(deleted_item_ids),
            vector_task_dispatched=needs_vector_sync,
        )

    def _read_workbook(self, uploaded_file):
        """读取上传文件并交给 openpyxl 解析。"""

        uploaded_file.seek(0)
        return load_workbook(uploaded_file, data_only=True)

    def _normalize_header(self, value) -> str:
        """把 Excel 表头统一成可匹配形式。"""

        return normalize_text(value).replace("（", "(").replace("）", ")")

    def _find_header_row(self, worksheet) -> int:
        """定位表头行。"""

        required_keys = {"material_name", "spec_model", "unit", "benchmark_price"}
        alias_map = {
            key: {self._normalize_header(alias) for alias in aliases}
            for key, aliases in HEADER_ALIASES.items()
        }

        for row_idx in range(1, min(worksheet.max_row, 30) + 1):
            row_values = [
                self._normalize_header(
                    worksheet.cell(row=row_idx, column=col_idx).value
                )
                for col_idx in range(1, worksheet.max_column + 1)
            ]
            matched_keys = set()
            for key, aliases in alias_map.items():
                if any(cell in aliases for cell in row_values):
                    matched_keys.add(key)
            if required_keys.issubset(matched_keys):
                return row_idx

        raise ValueError("Excel 表头不符合模板要求，请使用系统模板重新整理。")

    def _build_header_index(self, worksheet, header_row: int) -> dict[str, int]:
        """构造字段到列号的映射。"""

        row_values = {
            self._normalize_header(worksheet.cell(row=header_row, column=col_idx).value): col_idx
            for col_idx in range(1, worksheet.max_column + 1)
        }

        index: dict[str, int] = {}
        for key, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                column = row_values.get(self._normalize_header(alias))
                if column:
                    index[key] = column
                    break
        return index

    def _parse_price_bounds(self, worksheet, *, row_no: int, header_index: dict[str, int]):
        """解析区间价格，支持“单列区间”和“上下限两列”两种模板。"""

        if header_index.get("price_range"):
            range_value = worksheet.cell(
                row=row_no,
                column=header_index["price_range"],
            ).value
            return parse_price_range(range_value)

        low = parse_decimal(
            worksheet.cell(row=row_no, column=header_index["price_low"]).value
        ) if header_index.get("price_low") else None
        high = parse_decimal(
            worksheet.cell(row=row_no, column=header_index["price_high"]).value
        ) if header_index.get("price_high") else None
        return low, high

    def _build_embedding_text_from_row(self, row: ParsedGovernmentPriceRow) -> str:
        """根据一条标准价行构造用于向量检索的文本。"""

        return build_embedding_text(
            material_name=row.material_name_normalized,
            spec_model=row.spec_model_raw,
            unit=row.unit_raw,
        )

    def _build_row_key_from_row(self, row: ParsedGovernmentPriceRow) -> tuple[str, str, str]:
        """构造 Excel 行的业务唯一键。"""

        return (
            row.material_name_normalized,
            row.spec_model_normalized,
            row.unit_normalized,
        )

    def _build_row_key_from_item(self, item: GovernmentPriceItem) -> tuple[str, str, str]:
        """构造数据库明细行的业务唯一键。"""

        return (
            item.material_name_normalized,
            item.spec_model_normalized,
            item.unit_normalized,
        )


government_price_service = GovernmentPriceService()
