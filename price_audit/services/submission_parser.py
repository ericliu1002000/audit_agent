"""价格送审表解析服务。"""

from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from price_audit.models import PriceAuditSubmission, PriceAuditSubmissionRow
from price_audit.services.normalization import normalize_text, parse_decimal


SUMMARY_FEE_TYPES = {"小计", "税费", "合计"}
SUBMISSION_HEADER_ROW = 2
DATA_START_ROW = 3
SUBMISSION_TEMPLATE_MARKERS = {
    (1, 1): "序号",
    (1, 2): "费用类型",
    (1, 3): "送审",
    (1, 9): "审核",
    (1, 15): "审减原因/未审减原因",
    (2, 3): "计量单位",
    (2, 4): "单价（元）",
    (2, 5): "数量",
    (2, 6): "天数",
    (2, 7): "预算金额（元）",
    (2, 8): "预算编制说明",
}


@dataclass(frozen=True)
class ParsedSubmissionRow:
    """送审表中的结构化行数据。"""

    excel_row_no: int
    sequence_no: str
    parent_sequence_no: str
    row_type: str
    fee_type: str
    submitted_unit: str
    submitted_unit_price: Decimal | None
    submitted_quantity: Decimal | None
    submitted_days: Decimal | None
    submitted_amount: Decimal | None
    budget_note: str
    raw_row_data: dict[str, Any]


def _clean_header_text(value: Any) -> str:
    """清洗表头文本，兼容换行和空格差异。"""

    return normalize_text(value).replace(" ", "")


def _normalize_sequence(value: Any) -> str:
    """把 Excel 序号尽量稳定地转成字符串。"""

    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, "g")
    return normalize_text(value).replace(" ", "")


def _infer_parent_sequence(sequence_no: str) -> str:
    """根据点号序号推断父级序号。"""

    if "." not in sequence_no:
        return ""
    return sequence_no.rsplit(".", 1)[0]


def _validate_template(sheet) -> None:
    """校验是否符合当前支持的送审表模板。"""

    for (row_no, col_no), expected in SUBMISSION_TEMPLATE_MARKERS.items():
        actual = _clean_header_text(sheet.cell(row=row_no, column=col_no).value)
        if actual != _clean_header_text(expected):
            raise ValueError("价格审核模板格式错误，请上传当前支持的送审表模板。")


def parse_submission_excel(file_path: str) -> list[ParsedSubmissionRow]:
    """解析价格送审 Excel，输出结构化行列表。"""

    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"送审文件不存在: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError("仅支持上传 .xlsx 格式文件。")

    try:
        workbook = load_workbook(path, data_only=True)
    except (InvalidFileException, ValueError) as exc:
        raise ValueError("仅支持上传 .xlsx 格式文件。") from exc

    sheet = workbook.worksheets[0]
    _validate_template(sheet)

    preliminary_rows: list[dict[str, Any]] = []
    child_parent_map: dict[str, int] = {}

    for row_no in range(DATA_START_ROW, sheet.max_row + 1):
        sequence_no = _normalize_sequence(sheet.cell(row=row_no, column=1).value)
        fee_type = normalize_text(sheet.cell(row=row_no, column=2).value)
        submitted_unit = normalize_text(sheet.cell(row=row_no, column=3).value)
        submitted_unit_price = parse_decimal(sheet.cell(row=row_no, column=4).value)
        submitted_quantity = parse_decimal(sheet.cell(row=row_no, column=5).value)
        submitted_days = parse_decimal(sheet.cell(row=row_no, column=6).value)
        submitted_amount = parse_decimal(sheet.cell(row=row_no, column=7).value)
        budget_note = normalize_text(sheet.cell(row=row_no, column=8).value)

        if not any(
            [
                sequence_no,
                fee_type,
                submitted_unit,
                submitted_unit_price is not None,
                submitted_quantity is not None,
                submitted_days is not None,
                submitted_amount is not None,
                budget_note,
            ]
        ):
            continue

        parent_sequence_no = _infer_parent_sequence(sequence_no)
        if parent_sequence_no:
            child_parent_map[parent_sequence_no] = child_parent_map.get(parent_sequence_no, 0) + 1

        preliminary_rows.append(
            {
                "excel_row_no": row_no,
                "sequence_no": sequence_no,
                "parent_sequence_no": parent_sequence_no,
                "fee_type": fee_type,
                "submitted_unit": submitted_unit,
                "submitted_unit_price": submitted_unit_price,
                "submitted_quantity": submitted_quantity,
                "submitted_days": submitted_days,
                "submitted_amount": submitted_amount,
                "budget_note": budget_note,
                "raw_row_data": {
                    "sequence_no": sequence_no,
                    "fee_type": fee_type,
                    "submitted_unit": submitted_unit,
                    "submitted_unit_price": (
                        str(submitted_unit_price) if submitted_unit_price is not None else None
                    ),
                    "submitted_quantity": (
                        str(submitted_quantity) if submitted_quantity is not None else None
                    ),
                    "submitted_days": str(submitted_days) if submitted_days is not None else None,
                    "submitted_amount": (
                        str(submitted_amount) if submitted_amount is not None else None
                    ),
                    "budget_note": budget_note,
                },
            }
        )

    if not preliminary_rows:
        raise ValueError("送审表中未解析到有效数据。")

    rows: list[ParsedSubmissionRow] = []
    for item in preliminary_rows:
        fee_type = item["fee_type"]
        sequence_no = item["sequence_no"]
        if fee_type in SUMMARY_FEE_TYPES:
            row_type = PriceAuditSubmissionRow.RowType.SUMMARY
        elif sequence_no and child_parent_map.get(sequence_no):
            row_type = PriceAuditSubmissionRow.RowType.GROUP
        else:
            row_type = PriceAuditSubmissionRow.RowType.LEAF

        rows.append(
            ParsedSubmissionRow(
                excel_row_no=item["excel_row_no"],
                sequence_no=sequence_no,
                parent_sequence_no=item["parent_sequence_no"],
                row_type=row_type,
                fee_type=fee_type,
                submitted_unit=item["submitted_unit"],
                submitted_unit_price=item["submitted_unit_price"],
                submitted_quantity=item["submitted_quantity"],
                submitted_days=item["submitted_days"],
                submitted_amount=item["submitted_amount"],
                budget_note=item["budget_note"],
                raw_row_data=item["raw_row_data"],
            )
        )
    return rows


def populate_submission_rows(submission: PriceAuditSubmission) -> list[PriceAuditSubmissionRow]:
    """解析送审文件并写入送审明细行。"""

    parsed_rows = parse_submission_excel(submission.source_file.path)
    rows = [
        PriceAuditSubmissionRow(
            submission=submission,
            excel_row_no=item.excel_row_no,
            sequence_no=item.sequence_no,
            parent_sequence_no=item.parent_sequence_no,
            row_type=item.row_type,
            fee_type=item.fee_type,
            submitted_unit=item.submitted_unit,
            submitted_unit_price=item.submitted_unit_price,
            submitted_quantity=item.submitted_quantity,
            submitted_days=item.submitted_days,
            submitted_amount=item.submitted_amount,
            budget_note=item.budget_note,
            raw_row_data=item.raw_row_data,
        )
        for item in parsed_rows
    ]
    created_rows = PriceAuditSubmissionRow.objects.bulk_create(rows)
    return created_rows
