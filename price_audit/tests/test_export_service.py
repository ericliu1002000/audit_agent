"""价格审核 Excel 导出测试。"""

from __future__ import annotations

from io import BytesIO

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl import load_workbook

from price_audit.models import GovernmentPriceBatch
from price_audit.services.export_service import build_audited_excel_content
from price_audit.tests.helpers import (
    TempMediaRootMixin,
    build_price_audit_submission_workbook,
    create_row_decision,
    create_submission_row,
    create_submission_with_workbook,
)


User = get_user_model()


class ExportServiceTests(TempMediaRootMixin, TestCase):
    """验证审核表回填导出。"""

    def setUp(self):
        self.user = User.objects.create_user(username="export", password="Testpass123")
        self.batch = GovernmentPriceBatch.objects.create(
            region_name="天津",
            year=2026,
            is_active=True,
            vector_status=GovernmentPriceBatch.VectorStatus.ACTIVE,
        )

    def test_build_audited_excel_content_fills_missing_headers_and_decision_cells(self):
        """导出时应补全表头并回填审核结果。"""

        submission = create_submission_with_workbook(
            created_by=self.user,
            price_batch=self.batch,
            workbook=build_price_audit_submission_workbook(clear_audit_headers=True),
        )
        row = create_submission_row(submission, excel_row_no=3, sequence_no="1", fee_type="场地租")
        create_row_decision(
            row,
            reviewed_unit="平米",
            reviewed_unit_price="8.5000",
            reviewed_quantity="100.0000",
            reviewed_days="3.0000",
            reviewed_amount="2550.0000",
            reduction_amount="150.0000",
            reason="参考价格后下调。",
        )

        content = build_audited_excel_content(submission)
        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.worksheets[0]

        self.assertEqual(sheet.cell(row=1, column=9).value, "审核")
        self.assertEqual(sheet.cell(row=1, column=15).value, "审减原因/未审减原因")
        self.assertEqual(sheet.cell(row=2, column=13).value, "审核金额（元）")
        self.assertEqual(sheet.cell(row=3, column=10).value, 8.5)
        self.assertEqual(sheet.cell(row=3, column=13).value, 2550)
        self.assertEqual(sheet.cell(row=3, column=15).value, "参考价格后下调。")

    def test_build_audited_excel_content_skips_rows_without_decision(self):
        """没有审核结果的行不应被错误回填。"""

        submission = create_submission_with_workbook(
            created_by=self.user,
            price_batch=self.batch,
        )
        decided_row = create_submission_row(submission, excel_row_no=3, sequence_no="1", fee_type="场地租")
        undecided_row = create_submission_row(submission, excel_row_no=4, sequence_no="2", fee_type="电费")
        create_row_decision(decided_row, reviewed_amount="2700.0000")

        content = build_audited_excel_content(submission)
        workbook = load_workbook(BytesIO(content), data_only=True)
        sheet = workbook.worksheets[0]

        self.assertEqual(sheet.cell(row=3, column=13).value, 2700)
        self.assertIsNone(sheet.cell(row=4, column=13).value)
        self.assertEqual(undecided_row.fee_type, "电费")
