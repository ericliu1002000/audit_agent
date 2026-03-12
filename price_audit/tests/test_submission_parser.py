"""价格送审表解析测试。"""

from __future__ import annotations

from pathlib import Path
from tempfile import NamedTemporaryFile
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase
from openpyxl.utils.exceptions import InvalidFileException

from price_audit.models import GovernmentPriceBatch, PriceAuditSubmission
from price_audit.services.submission_parser import parse_submission_excel, populate_submission_rows
from price_audit.tests.helpers import (
    TempMediaRootMixin,
    build_price_audit_submission_workbook,
    create_submission_with_workbook,
)


User = get_user_model()


class SubmissionParserTests(TempMediaRootMixin, TestCase):
    """验证送审表解析逻辑。"""

    def setUp(self):
        self.user = User.objects.create_user(username="parser", password="Testpass123")
        self.batch = GovernmentPriceBatch.objects.create(
            region_name="天津",
            year=2026,
            is_active=True,
            vector_status=GovernmentPriceBatch.VectorStatus.ACTIVE,
        )

    def _build_submission(self) -> PriceAuditSubmission:
        return create_submission_with_workbook(
            created_by=self.user,
            price_batch=self.batch,
        )

    def test_parse_submission_excel_classifies_row_types(self):
        """解析结果应正确识别 leaf/group/summary。"""

        submission = self._build_submission()
        rows = parse_submission_excel(submission.source_file.path)

        by_sequence = {row.sequence_no or row.fee_type: row for row in rows}
        self.assertEqual(by_sequence["1"].row_type, "leaf")
        self.assertEqual(by_sequence["2"].row_type, "group")
        self.assertEqual(by_sequence["2.1"].parent_sequence_no, "2")
        self.assertEqual(by_sequence["3"].row_type, "leaf")
        self.assertEqual(by_sequence["税费"].row_type, "summary")

    def test_parse_submission_excel_skips_blank_rows_and_keeps_leaf_amount_only_rows(self):
        """空白行应跳过，只有金额和说明的行仍是 leaf。"""

        workbook = build_price_audit_submission_workbook(
            data_rows=[
                [1, "场地租", "平米", 9, 100, 3, 2700, "场地租赁说明", None, None, None, None, None, None, None],
                [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
                [2, "电费", None, None, None, None, 500, "电费预估", None, None, None, None, None, None, None],
                [None, "税费", "-", "-", "-", "-", 100, "税费说明", None, None, None, None, None, None, None],
            ]
        )
        submission = create_submission_with_workbook(
            created_by=self.user,
            price_batch=self.batch,
            workbook=workbook,
        )

        rows = parse_submission_excel(submission.source_file.path)

        self.assertEqual(len(rows), 3)
        electric_fee = next(item for item in rows if item.fee_type == "电费")
        self.assertEqual(electric_fee.row_type, "leaf")
        self.assertEqual(electric_fee.sequence_no, "2")

    def test_parse_submission_excel_handles_dotted_parent_sequences(self):
        """点号序号应正确推断父级。"""

        workbook = build_price_audit_submission_workbook(
            data_rows=[
                ["2", "特装展台搭建", "-", None, None, None, 1500, "父项", None, None, None, None, None, None, None],
                ["2.1", "子项A", "㎡", 10, 2, None, 20, "A", None, None, None, None, None, None, None],
                ["2.10", "子项B", "㎡", 30, 2, None, 60, "B", None, None, None, None, None, None, None],
                ["11", "宣传", "-", None, None, None, 200, "宣传父项", None, None, None, None, None, None, None],
                ["11.2", "宣传子项", "项", 1, 1, None, 50, "宣传子项说明", None, None, None, None, None, None, None],
                [None, "小计", "-", "-", "-", "-", 1810, "-", None, None, None, None, None, None, None],
            ]
        )
        submission = create_submission_with_workbook(
            created_by=self.user,
            price_batch=self.batch,
            workbook=workbook,
        )

        rows = parse_submission_excel(submission.source_file.path)
        by_sequence = {row.sequence_no or row.fee_type: row for row in rows}

        self.assertEqual(by_sequence["2"].row_type, "group")
        self.assertEqual(by_sequence["2.10"].parent_sequence_no, "2")
        self.assertEqual(by_sequence["11"].row_type, "group")
        self.assertEqual(by_sequence["11.2"].parent_sequence_no, "11")

    def test_populate_submission_rows_persists_rows(self):
        """解析结果应落库到 SubmissionRow。"""

        submission = self._build_submission()
        rows = populate_submission_rows(submission)

        self.assertEqual(len(rows), 8)
        persisted = submission.rows.get(sequence_no="3")
        self.assertEqual(persisted.fee_type, "电费")
        self.assertEqual(persisted.row_type, "leaf")
        self.assertEqual(str(persisted.submitted_amount), "500.0000")

    def test_parse_submission_excel_rejects_invalid_template(self):
        """非当前模板应报业务错误。"""

        path = Path(self._build_submission().source_file.path)
        broken = path.with_name("invalid.xlsx")
        broken.write_bytes(path.read_bytes())
        from openpyxl import load_workbook

        workbook = load_workbook(broken)
        workbook.active["A1"] = "错误表头"
        workbook.save(broken)

        with self.assertRaisesRegex(ValueError, "价格审核模板格式错误"):
            parse_submission_excel(str(broken))

    def test_parse_submission_excel_rejects_missing_file(self):
        """文件不存在时应直接报错。"""

        with self.assertRaises(FileNotFoundError):
            parse_submission_excel("/tmp/not-found-price-audit.xlsx")

    def test_parse_submission_excel_rejects_non_xlsx_suffix(self):
        """非 xlsx 扩展名应拒绝解析。"""

        with NamedTemporaryFile(suffix=".csv") as temp_file:
            temp_file.write(b"bad")
            temp_file.flush()
            with self.assertRaisesRegex(ValueError, "仅支持上传 .xlsx 格式文件"):
                parse_submission_excel(temp_file.name)

    @patch("price_audit.services.submission_parser.load_workbook")
    def test_parse_submission_excel_rejects_invalid_workbook_content(self, workbook_mock):
        """工作簿无法读取时应返回统一业务错误。"""

        workbook_mock.side_effect = InvalidFileException("bad workbook")
        with NamedTemporaryFile(suffix=".xlsx") as temp_file:
            temp_file.write(b"bad")
            temp_file.flush()
            with self.assertRaisesRegex(ValueError, "仅支持上传 .xlsx 格式文件"):
                parse_submission_excel(temp_file.name)

    def test_parse_submission_excel_rejects_when_no_valid_data_rows(self):
        """没有有效数据行时应报错。"""

        workbook = build_price_audit_submission_workbook(data_rows=[])
        with NamedTemporaryFile(suffix=".xlsx") as temp_file:
            temp_file.write(workbook.read())
            temp_file.flush()
            with self.assertRaisesRegex(ValueError, "送审表中未解析到有效数据"):
                parse_submission_excel(temp_file.name)

    def test_parse_submission_excel_supports_sample_template_file(self):
        """仓库中的样例模板应能成功解析。"""

        sample_path = (
            Path(__file__).resolve().parents[2] / "docs" / "price_audit" / "评审表样例.xlsx"
        )
        rows = parse_submission_excel(str(sample_path))

        self.assertGreater(len(rows), 0)
        self.assertTrue(any(item.fee_type == "税费" and item.row_type == "summary" for item in rows))
